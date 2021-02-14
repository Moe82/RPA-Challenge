from selenium import webdriver
from bs4 import BeautifulSoup as bs
import requests
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import selenium.webdriver
import pandas as pd
from time import sleep
import os
from openpyxl import load_workbook, Workbook

__location__ = os.path.realpath(os.path.join(os.getcwd(), os.path.dirname(__file__))) # path to current dir

def get_investment_table(agenecy_summary_url):
	""" extracts the investment table from the agency's summary page and returns it as a Panads DF object """
	driver = webdriver.Chrome(executable_path=os.path.join(__location__, 'chromedriver'))
	driver.get(agenecy_summary_url)
	WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.ID,"investments-table-object")))
	driver.find_element_by_xpath('//*[@id="investments-table-object_length"]/label/select/option[4]').click()
	sleep(10)  # not good practice, find more optimal solution later. 
	investment_table_df = pd.read_html(driver.page_source, attrs={'id': 'investments-table-object'})[0]
	driver.close()
	driver.quit()
	return(investment_table_df)

def get_agency_spendings():
	""" extracts the names of all agenecies and thier yearly spendings and returns the data as a Panads DF object """
	driver = webdriver.Chrome(executable_path=os.path.join(__location__, 'chromedriver'))
	driver.get('https://itdashboard.gov/')
	driver.find_element_by_xpath('//*[@id="node-23"]/div/div/div/div/div/div/div/a').click()
	WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.CSS_SELECTOR,"#agency-tiles-widget")))
	agency_tiles_widget_list = driver.find_element_by_css_selector('#agency-tiles-widget').text.splitlines()
	agency_tiles_widget_list = [agency_tiles_widget_list[i:i+4] for i in range(0, len(agency_tiles_widget_list), 4)] # convert 1d list into list of lists
	for list in agency_tiles_widget_list:
		""" pop every 2nd and 3rd element to create sublists containing only the agenecy name and total spending """ 
		list.pop(1)
		list.pop(2)
	agency_tiles_widget_df = pd.DataFrame(agency_tiles_widget_list)
	agency_tiles_widget_df.columns = ["Agenecy", "Total FY2020 Spending"] 
	driver.close()
	driver.quit()
	return (agency_tiles_widget_df)


if __name__ == "__main__":
	Workbook().save(os.path.join(__location__, 'output/output.xlsx'))
	writer = pd.ExcelWriter(os.path.join(__location__, 'output/output.xlsx'), engine = 'openpyxl')
	book = load_workbook(os.path.join(__location__, 'output/output.xlsx'))
	writer.book = book
	writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
	agenecy_spendings = get_agency_spendings()
	agenecy_spendings.to_excel(writer, header=True, index=False, sheet_name="Agencies")
	with open('investments_to_scrape') as f:
		content = f.readlines()
	content = [x.strip() for x in content]
	for url in content:
		investment_table = get_investment_table(url)
		investment_table.to_excel(writer, sheet_name=url[-3:])
	writer.save()
	